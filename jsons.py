import csv
import json
import logging
import os

from openpyxl import Workbook, load_workbook
from time import perf_counter as pc
from api_hh_connect import APIHHConnect
from dirs import Dirs
from s3 import S3


class JSONs:
    @staticmethod
    def save_vacancies_json(vacancies, params):
        page = "" if params['page'] is None else f"---{params['page']}"
        host_path = f"C:/Users/mi/work/hh_parser/group_jsons/{params['date_from'][:13]}{page}.json"
        s3_path = "group_jsons/" + f"{params['date_from'][:13]}{page}.json"
        if os.path.exists(host_path):
            return
        with open(host_path, 'w', encoding='utf8') as outfile:
            json.dump(vacancies, outfile, sort_keys=False, indent=4, ensure_ascii=False, separators=(',', ': '))
        S3.upload(host_path, s3_path)

    @staticmethod
    def upload_all_jsons():
        path_from = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/json_urls/'
        filename = '2022-11-15.xlsx'
        path_to = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/vacancies_jsons/'
        start = pc()
        count = 0
        # for filename in os.listdir(path_from):
        directory = filename[:-5]
        Dirs.make_dir(path_to, directory)
        wb = load_workbook(path_from + filename)
        ws = wb.active
        row_number = 0
        logging.warning(f"{filename} started uploading, {len(list(ws.values))} JSONs founded")
        for row in ws.values:
            row_number += 1
            if row[2]:
                continue
            logging.info(row)
            count += 1
            logging.info(row[1] + ' ' + str(count) + '   ' + str(pc() - start))
            try:
                JSONs.threads_upload_jsons(row[1], directory, row[0])
                ws.cell(row=row_number, column=3).value = True
            except Exception as e:
                ws.cell(row=row_number, column=3).value = False
                logging.error(e)
        logging.warning(f"{filename} uploaded")
        wb.save(path_from + filename)

    @staticmethod
    def threads_upload_jsons(url, directory, vacancy_id):
        host_path = f'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/vacancies_jsons/{directory}/{vacancy_id}.json'
        s3_path = f'vacancies_jsons/{directory}/{vacancy_id}.json'
        if not os.path.exists(host_path):
            vacancy = APIHHConnect.connect(url)
            with open(host_path, 'w', encoding='utf8') as outfile:
                json.dump(vacancy, outfile, sort_keys=False, indent=4, ensure_ascii=False, separators=(',', ': '))
        S3.upload(host_path, s3_path)
