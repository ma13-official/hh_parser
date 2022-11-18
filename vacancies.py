import csv
import json
import os
import threading

from api_hh_connect import APIHHConnect
from openpyxl import Workbook, load_workbook
import datetime
from time import perf_counter as pc
from vacancy import Vacancy
from yadisk_connect import YaDisk


class Vacancies:
    name_of_fields = ['id', 'name', 'area', 'salary', 'address', 'response_url', 'published_at', 'created_at',
                      'archived', 'apply_alternate_url', 'alternate_url', 'employer', 'snippet', 'schedule',
                      'working_days', 'working_time_intervals', 'working_time_modes', 'accept_temporary']
    vacancies_array = {}
    currencies = {}
    number_of_vacancies, number_of_changes, every_day_checks_counter, every_hour_checks_counter, count_duplicates = \
        [0 for _ in range(5)]
    json_urls = []
    errors = 0

    @staticmethod
    def get_currencies():
        currencies = {}
        dictionaries = APIHHConnect.connect('dictionaries')
        for currency in dictionaries['currency']:
            currencies[currency['code']] = currency['rate']
        return currencies

    def check_all(self, days):
        query = "vacancies"
        now = datetime.date.today()  # сегодняшняя дата
        today = datetime.datetime(now.year, now.month, now.day, 0, 0)  # сегодня в 00:00:00
        date_from = (today - datetime.timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%S")
        date_to = today.strftime("%Y-%m-%dT%H:%M:%S")
        params = {'specialization': '1', 'date_from': date_from, 'date_to': date_to, 'per_page': 100}
        vacancies = APIHHConnect.connect(query, params)
        print(vacancies['found'])
        if vacancies['found'] > 2000:
            self.separating_by_days(query, params, today, days)
        else:
            print("!!!!!!!!!!!")
            print("!!!ERROR!!!")
            print("!!!ERROR!!!")
            print("!!!ERROR!!!")
            print("!!!!!!!!!!!")

    def separating_by_days(self, query, params, today, days):
        for days_ago in range(days):
            self.vacancies_array = {}
            params['date_to'] = (today - datetime.timedelta(days=days_ago)).strftime("%Y-%m-%dT%H:%M:%S")
            params['date_from'] = (today - datetime.timedelta(days=days_ago + 1)).strftime("%Y-%m-%dT%H:%M:%S")
            params['page'] = None
            vacancies = APIHHConnect.connect(query, params)
            print(params)
            print("По дням:" + str(vacancies['found']))
            self.every_day_checks_counter += vacancies['found']
            self.make_dir('C:/Users/mi/work/hh_parser/vacancies/',
                          (today - datetime.timedelta(days=days_ago + 1)).strftime("%Y-%m-%d"))
            if vacancies['found'] > 2000:
                # self.separating_by_hours(query, params, today, days_ago)
                self.separating_by_hours_with_threads(query, today, days_ago)
            else:
                print(query, params)
                self.check_pages(query, params)
                print(len(list(self.vacancies_array.keys())))
                self.number_of_changes += (len(list(self.vacancies_array.keys())))
            date = (today - datetime.timedelta(days=days_ago + 1)).strftime("%Y-%m-%d")
            self.write_in_csv(self.vacancies_array, date)
            self.write_jsons_in_csv(self.json_urls, date)
            self.json_urls = []

    @staticmethod
    def make_dir(path, directory):
        os.chdir(path)
        if not os.path.exists(os.path.join(path, directory)):
            os.mkdir(directory)

    def separating_by_hours(self, query, params, today, days_ago):
        cur_day = today - datetime.timedelta(days=days_ago)
        self.query_between_hours(query, params, cur_day, 0, 4)
        for hour in range(4, 18):
            self.query_between_hours(query, params, cur_day, hour, hour+1)
        self.query_between_hours(query, params, cur_day, 18, 20)
        self.query_between_hours(query, params, cur_day, 20, 24)

    def separating_by_hours_with_threads(self, query, today, days_ago):
        cur_day = today - datetime.timedelta(days=days_ago)
        for hour in range(0, 24, 4):
            self.threads(query, cur_day, hour)

    def threads(self, query,  cur_day, hour):
        threads = []
        for x in range(4):
            params = {'specialization': '1',
                      'date_from': (cur_day - datetime.timedelta(hours=hour + x + 1)).strftime("%Y-%m-%dT%H:%M:%S"),
                      'date_to': (cur_day - datetime.timedelta(hours=hour + x)).strftime("%Y-%m-%dT%H:%M:%S"),
                      'per_page': 100}
            t = threading.Thread(target=self.query_between_hours_with_threads, args=(query, params))
            threads.append(t)
            t.start()
            params = {}
        for t in threads:
            t.join()

    def query_between_hours_with_threads(self, query, params):
        vacancies = APIHHConnect.connect(query, params)
        print(params)
        print("По часам:" + str(vacancies['found']))
        self.end_of_qbh(vacancies, query, params)

    def query_between_hours(self, query, params, cur_day, date_to_hours, date_from_hours):
        params['date_to'] = (cur_day - datetime.timedelta(hours=date_to_hours)).strftime("%Y-%m-%dT%H:%M:%S")
        params['date_from'] = (cur_day - datetime.timedelta(hours=date_from_hours)).strftime("%Y-%m-%dT%H:%M:%S")
        params['page'] = None
        # print(params)

        vacancies = APIHHConnect.connect(query, params)
        print(params, date_to_hours, date_from_hours)
        print("По часам:" + str(vacancies['found']))
        self.end_of_qbh(vacancies, query, params)

    def end_of_qbh(self, vacancies, query, params):
        self.every_hour_checks_counter += vacancies['found']
        if vacancies['found'] > 2000:
            print("!!!!!!!!!!!")
            print("!!!ERROR!!!")
            print("!!!ERROR!!!")
            print("!!!ERROR!!!")
            print("!!!!!!!!!!!")
        cur_len = len(list(self.vacancies_array.keys()))
        self.check_pages(query, params)
        self.number_of_changes += (len(self.vacancies_array) - cur_len)

    def check_pages(self, query, params):
        # print(params)
        params['page'] = None
        vacancies = APIHHConnect.connect(query, params)
        pages = vacancies['found'] // 100
        if pages > 0:
            for i in range(pages + 1):
                params['page'] = i
                vacancies = APIHHConnect.connect(query, params)
                self.get_needed_fields(vacancies)
        else:
            self.get_needed_fields(vacancies)

    def get_needed_fields(self, vacancies):
        """
        :param vacancies: JSON-объект полученный после запроса
        :return: заполнение двумерного массива vacancies_array необходимыми полями.
        """
        for vacancy in vacancies[list(vacancies.keys())[0]]:
            self.number_of_vacancies += 1
            if vacancy['archived']:
                continue
            vacancy_to_add = Vacancy.create_vacancy(vacancy, self.currencies)
            if vacancy_to_add not in self.vacancies_array.values():
                self.vacancies_array[vacancy_to_add.vacancy_id] = vacancy_to_add
                self.json_urls.append([vacancy_to_add.vacancy_id, vacancy_to_add.url_json])
            else:
                if not Vacancy.check_csv(vacancy_to_add):
                    self.count_duplicates += 1

    @staticmethod
    def write_in_csv(vacancies_array, date):
        """
        :return: запись двумерного массива в Excel-таблицу.
        """
        path = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/vacancies_per_day/' + date + '.xlsx'
        name_of_fields = ['number', 'vacancy_id', 'name', 'area', 'address', 'employer', 'url_hh', 'url_json',
                          'published_at', 'created_at', 'salary_from', 'salary_to', 'requirement', 'responsibility',
                          'schedule', 'working_days', 'working_time_intervals', 'working_time_modes',
                          'accept_temporary']
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(name_of_fields)
        number = 0
        for item in vacancies_array.values():
            number += 1
            row = [number]
            for value in item.__dict__.values():
                row.append(value)
            ws.append(row)
        wb.save(path)

    @staticmethod
    def write_jsons_in_csv(json_urls, date):
        path = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/json_urls/' + date + '.xlsx'
        name_of_fields = ['number', 'vacancy_id', 'name', 'area', 'address', 'employer', 'url_hh', 'url_json',
                          'published_at', 'created_at', 'salary_from', 'salary_to', 'requirement', 'responsibility',
                          'schedule', 'working_days', 'working_time_intervals', 'working_time_modes',
                          'accept_temporary']
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
        ws = wb.active
        for json_url in json_urls:
            ws.append(json_url)
        wb.save(path)

    @staticmethod
    def upload_all_jsons():
        path = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/json_urls/'
        filename = '2022-11-15.csv'
        # for filename in os.listdir(path):
        cur = open(path + filename)
        directory = filename[:-4]
        Vacancies.make_dir('C:/Users/mi/work/hh_parser/vacancies_jsons/', directory)
        file_reader = csv.DictReader(cur, delimiter=",")
        print(cur)
        count = 0
        start = pc()
        for row in file_reader:
            print(row)
            count += 1
            print(row['2'] + ' ' + str(count) + '   ' + str(pc()-start))
            Vacancies.threads_upload_jsons(row['1'], directory)

    @staticmethod
    def threads_upload_jsons(url, directory):
        vacancy = APIHHConnect.connect(url)
        with open('C:/Users/mi/work/hh_parser/vacancies_jsons/' + directory + '/' + vacancy['id'] + '.json', 'w',
                  encoding='utf8') as outfile:
            try:
                json.dump(vacancy, outfile, sort_keys=False, indent=4, ensure_ascii=False, separators=(',', ': '))
            except Exception as e:
                print(e)

    def start(self):
        """
        :return: запуск приложения
        """
        self.currencies = self.get_currencies()
        self.check_all(30)
        # YaDisk.write_on_yadisk()
        print(self.every_day_checks_counter)
        print(self.every_hour_checks_counter)
        print(self.count_duplicates)
        # self.upload_all_jsons()
        # print(self.errors)


start = pc()
Vacancies().start()
print(pc() - start)
