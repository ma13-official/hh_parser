import logging
import os
from openpyxl import Workbook, load_workbook


class CSVs:
    @staticmethod
    def write_in_csv(vacancies_array, date):
        """
        :return: запись двумерного массива в Excel-таблицу.
        """
        path = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/vacancies_per_day/' + date + '.xlsx'
        name_of_fields = True
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            name_of_fields = False
        else:
            wb = Workbook()
            ws = wb.active
        number = 0
        for item in vacancies_array.values():
            if number == 0 and name_of_fields:
                ws.append(['number'] + list(item.__dict__.keys()))
            number += 1
            row = [number]
            for value in item.__dict__.values():
                row.append(value)
            ws.append(row)
        wb.save(path)

    @staticmethod
    def create_csv(vacancy):
        PATH = 'C:/Users/mi/work/hh_parser/vacancies/'
        wb = Workbook()
        ws = wb.active
        ws.append(['number'] + list(vacancy.__dict__.keys()))
        row = []
        for value in vacancy.__dict__.values():
            row.append(value)
        ws.append(row)
        date = vacancy.created_at[:10]
        wb.save(PATH + date + '/' + str(vacancy.vacancy_id) + '.csv')

    @staticmethod
    def check_len_csv(date, founded, first_check):
        path = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/vacancies_per_day/' + date + '.xlsx'
        wb = load_workbook(path)
        ws = wb.active
        wb_len = len(list(ws.values))
        if first_check:
            if wb_len == founded:
                logging.info("The day was checked for the first time and all vacancies were found")
            else:
                logging.error(f"The day was checked for the first time and {founded - wb_len} vacancies were not found")
        else:
            if wb_len == founded:
                logging.info("The day was checked not for the first time, no changes")
            elif wb_len > founded:
                logging.info(f"There are {wb_len - founded} more vacancies in the dataset")
            else:
                logging.error(
                    f"The day was checked not for the first time and {founded - wb_len} vacancies were not found")

    @staticmethod
    def write_jsons_in_csv(json_urls, date):
        path = 'C:/Users/mi/OneDrive - ITMO UNIVERSITY/work/hh_parser/json_urls/' + date + '.xlsx'
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
        ws = wb.active
        for json_url in json_urls:
            ws.append(json_url)
        wb.save(path)

    @staticmethod
    def check_csv(vacancy):
        return not os.path.exists('C:/Users/mi/work/hh_parser/vacancies/' + vacancy.created_at[:10] + '/'
                                  + str(vacancy.vacancy_id) + '.csv')
